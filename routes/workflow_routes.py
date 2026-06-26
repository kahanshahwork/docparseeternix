"""
routes/workflow_routes.py — Modules 2/3/6/7/8/9 (HTTP layer).

This is the ONLY file that should change for approval-screen, category-
allocation, GST-review, or P&L-summary behavior. It calls into
core/category_master.py, core/gst_engine.py, core/pnl_engine.py and
core/vendor_memory.py — edit THOSE for logic changes, this file for
endpoint/request-shape changes only.

Flow: POST /api/statements (ingest parsed txns)
      -> GET/PATCH /api/statements/<id>/transactions (approval screen, item 5)
      -> POST /api/statements/<id>/approve
      -> GET /api/categories (item 6)
      -> GET /api/statements/<id>/groups (item 7)
      -> POST /api/statements/<id>/categorize (bulk or single)
      -> GET /api/statements/<id>/gst (item 8, editable table)
      -> PATCH /api/transactions/<id> (live amount edit -> GST recalced, item 8)
      -> GET /api/statements/<id>/pnl (item 9)
"""

from flask import Blueprint, request, jsonify
from core.db import get_db, log_audit
from core import category_master, gst_engine, pnl_engine, vendor_memory, category_engine
from core.business_types import is_valid_business_type

workflow_bp = Blueprint("workflow", __name__, url_prefix="/api")


def _row_to_txn(row) -> dict:
    return dict(row)


def get_client_id_for_statement(conn, statement_id: int):
    """
    Resolves a statement's owning client_id via statements.quarter_id ->
    quarters.client_id. Returns None if the statement has no quarter_id set
    (e.g. an old statement created before the client/quarter selector was
    added to the Upload & Parse page) -- callers should treat None as "no
    client context available" rather than guessing.
    """
    row = conn.execute(
        """SELECT q.client_id FROM statements s
           JOIN quarters q ON q.id = s.quarter_id
           WHERE s.id = ?""",
        (statement_id,),
    ).fetchone()
    return row["client_id"] if row else None


def _hydrate_category_fields(conn, txn: dict) -> dict:
    """Attach category_name/pnl_group/bas_label/gst_rate for engine consumption."""
    if txn.get("category_id"):
        cat = conn.execute("SELECT * FROM categories WHERE id = ?", (txn["category_id"],)).fetchone()
        if cat:
            txn["category_name"] = cat["name"]
            txn["pnl_group"] = cat["pnl_group"]
            txn["bas_label"] = cat["bas_label"]
            txn["gst_rate"] = cat["gst_rate"]
            txn["gst_applicable"] = bool(cat["gst_applicable"])
            return txn
    txn["category_name"] = "Uncategorized"
    txn["pnl_group"] = "Excluded"
    txn["bas_label"] = "excluded"
    txn["gst_rate"] = 0.0
    txn["gst_applicable"] = False
    return txn


# ── Categories (item 6) ─────────────────────────────────────────────────────

@workflow_bp.route("/categories", methods=["GET"])
def get_categories():
    return jsonify(category_master.list_categories())


@workflow_bp.route("/categories", methods=["POST"])
def add_category():
    b = request.json or {}
    cat_id = category_master.create_category(
        code=b["code"], name=b["name"], pnl_group=b["pnl_group"],
        gst_applicable=b.get("gst_applicable", False),
        gst_rate=b.get("gst_rate", 0.10), bas_label=b.get("bas_label", "G11"),
    )
    return jsonify({"id": cat_id})


# ── Statement ingest (parsed -> stored, item 5) ─────────────────────────────

@workflow_bp.route("/statements", methods=["POST"])
def create_statement():
    """Takes the JSON output of /parse and stores it as a reviewable statement."""
    b = request.json or {}
    transactions = b.get("transactions", [])
    bank_id = b.get("bank_id", "unknown")
    filename = b.get("filename", "")
    quarter_id = b.get("quarter_id")

    statement_name = b.get("statement_name") or filename or None

    conn = get_db()
    cur = conn.execute(
        "INSERT INTO statements (quarter_id, bank_id, filename, statement_name, status) VALUES (?,?,?,?, 'parsed')",
        (quarter_id, bank_id, filename, statement_name),
    )
    statement_id = cur.lastrowid

    for t in transactions:
        group_key = vendor_memory.normalize_description(t.get("description", ""))
        conn.execute(
            """INSERT INTO transactions
               (statement_id, transaction_id, date, description, amount, balance,
                source_page, row_top, confidence, group_key)
               VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (statement_id, t.get("transaction_id"), t.get("date"), t.get("description"),
             t.get("amount", 0), t.get("balance"), t.get("source_page"),
             t.get("row_top", 0), t.get("confidence"), group_key),
        )
    conn.commit()
    return jsonify({"statement_id": statement_id, "count": len(transactions)})


# ── Approval screen (item 5) ────────────────────────────────────────────────

@workflow_bp.route("/statements/<int:sid>/transactions", methods=["GET"])
def list_statement_transactions(sid):
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM transactions WHERE statement_id = ? ORDER BY source_page, row_top", (sid,)
    ).fetchall()
    return jsonify([_row_to_txn(r) for r in rows])


@workflow_bp.route("/transactions/<int:tid>", methods=["PATCH"])
def update_transaction(tid):
    """Edit a single transaction (date/description/amount) and/or approve it.
    If amount changes and the row already has a category, GST is recalculated live (item 8)."""
    b = request.json or {}
    conn = get_db()
    txn = conn.execute("SELECT * FROM transactions WHERE id = ?", (tid,)).fetchone()
    if not txn:
        return jsonify({"error": "Not found"}), 404

    fields, vals = [], []
    for k in ("date", "description", "amount", "approved", "category_id"):
        if k in b:
            fields.append(f"{k} = ?")
            vals.append(b[k])

    new_amount = b.get("amount", txn["amount"])
    new_category_id = b.get("category_id", txn["category_id"])

    if new_category_id:
        cat = category_master.get_category(new_category_id)
        gst = gst_engine.recalc_transaction_gst(new_amount, cat)
        fields += ["gst_amount = ?", "net_amount = ?"]
        vals += [gst["gst_amount"], gst["net_amount"]]

    if fields:
        vals.append(tid)
        conn.execute(f"UPDATE transactions SET {', '.join(fields)} WHERE id = ?", vals)
        conn.commit()
        log_audit("transaction", tid, "edit", str(b))

    row = conn.execute("SELECT * FROM transactions WHERE id = ?", (tid,)).fetchone()
    return jsonify(_row_to_txn(row))


@workflow_bp.route("/statements/<int:sid>/approve", methods=["POST"])
def approve_statement(sid):
    """Bulk-approve all rows still pending, then advance statement status."""
    conn = get_db()
    conn.execute("UPDATE transactions SET approved = 1 WHERE statement_id = ?", (sid,))
    conn.execute("UPDATE statements SET status = 'approved' WHERE id = ?", (sid,))
    conn.commit()
    log_audit("statement", sid, "approve")
    return jsonify({"status": "approved"})


# ── Grouping for bulk categorization (item 7) ───────────────────────────────

@workflow_bp.route("/statements/<int:sid>/groups", methods=["GET"])
def get_groups(sid):
    """
    Returns groups split into two top-level buckets -- debit (money out)
    and credit (money in) -- since mixing the two in one flat list made no
    structural sense (an expense vendor and an income source should never
    be candidates for the same category). Within each bucket, transactions
    are grouped semantically as before (vendor_memory.group_transactions).

    Each group now includes the FULL transaction list (not just ids), so
    the frontend can render an expand/collapse view without extra round
    trips, and a `dominant_category_id` so the group's category dropdown
    correctly reflects what's already been assigned instead of always
    showing "Uncategorized" even after a category was applied.
    """
    conn = get_db()
    rows = [dict(r) for r in conn.execute(
        "SELECT * FROM transactions WHERE statement_id = ?", (sid,)
    ).fetchall()]

    debit_rows = [r for r in rows if r["amount"] < 0]
    credit_rows = [r for r in rows if r["amount"] >= 0]

    def build_groups(direction_rows):
        groups = vendor_memory.group_transactions(direction_rows)
        out = []
        for k, v in groups.items():
            cat_ids = {t["category_id"] for t in v if t["category_id"] is not None}
            dominant = next(iter(cat_ids)) if len(cat_ids) == 1 else None
            out.append({
                "group_key": k,
                "count": len(v),
                "total": round(sum(t["amount"] for t in v), 2),
                "sample_description": v[0]["description"],
                "dominant_category_id": dominant,
                "transactions": [
                    {
                        "id": t["id"], "date": t.get("date"), "description": t["description"],
                        "amount": t["amount"], "category_id": t["category_id"],
                    }
                    for t in v
                ],
            })
        out.sort(key=lambda g: -g["count"])
        return out

    return jsonify({
        "debit": build_groups(debit_rows),
        "credit": build_groups(credit_rows),
    })


@workflow_bp.route("/statements/<int:sid>/categorize/ai-batch", methods=["POST"])
def categorize_ai_batch(sid):
    """
    Runs the full two-stage engine (vendor memory -> semantic bucket -> ONE
    batched Groq call for everything left) across every uncategorized
    transaction in this statement. Returns suggestions only -- nothing is
    written to the database here, matching the "AI suggestions are never
    auto-applied" rule. The frontend shows these as suggestion badges the
    user accepts individually or in bulk via the normal /categorize endpoint.
    """
    conn = get_db()
    client_id = get_client_id_for_statement(conn, sid)
    if client_id is None:
        return jsonify({"error": "This statement has no client linked (no quarter_id) -- "
                                  "cannot resolve business_type for AI categorization."}), 400

    client_row = conn.execute("SELECT business_type FROM clients WHERE id = ?", (client_id,)).fetchone()
    business_type_code = client_row["business_type"] if client_row else "RETAIL_TRADING"

    rows = conn.execute(
        "SELECT * FROM transactions WHERE statement_id = ? AND category_id IS NULL", (sid,)
    ).fetchall()
    if not rows:
        return jsonify({})

    batch_input = [
        {
            "id": r["id"],
            "description": r["description"],
            "amount": r["amount"],
            "direction": "debit" if r["amount"] < 0 else "credit",
        }
        for r in rows
    ]

    results = category_engine.categorize_transactions_batch(client_id, batch_input, business_type_code)

    return jsonify({
        str(tid): {
            "category_id": s.category_id,
            "category_name": s.category_name,
            "confidence": s.confidence,
            "source": s.source,
            "gst_note": s.gst_note,
        }
        for tid, s in results.items()
    })


# ── Categorization (items 6, 7) ─────────────────────────────────────────────

@workflow_bp.route("/statements/<int:sid>/categorize", methods=["POST"])
def categorize(sid):
    """body: {transaction_ids: [..], category_id: N}  -- works for single or bulk/group assignment.
    client_id is resolved server-side from the statement's quarter -> client chain,
    NOT trusted from the request body, so vendor memory is always attributed to the
    correct client even if the frontend forgets to pass one."""
    b = request.json or {}
    tids = b.get("transaction_ids", [])
    category_id = b.get("category_id")  # None = un-categorize

    conn = get_db()
    client_id = get_client_id_for_statement(conn, sid)

    if category_id is None:
        # Un-categorize: clear category, gst_amount, net_amount
        for tid in tids:
            row = conn.execute("SELECT * FROM transactions WHERE id = ?", (tid,)).fetchone()
            if not row:
                continue
            conn.execute(
                "UPDATE transactions SET category_id = NULL, gst_amount = 0, net_amount = ? WHERE id = ?",
                (row["amount"], tid),
            )
        conn.commit()
        log_audit("statement", sid, "categorize", f"{len(tids)} txns -> Uncategorized")
        return jsonify({"updated": len(tids), "client_id": client_id})

    cat = category_master.get_category(category_id)
    if not cat:
        return jsonify({"error": "Unknown category"}), 400

    for tid in tids:
        row = conn.execute("SELECT * FROM transactions WHERE id = ?", (tid,)).fetchone()
        if not row:
            continue
        gst = gst_engine.recalc_transaction_gst(row["amount"], cat)
        conn.execute(
            "UPDATE transactions SET category_id = ?, gst_amount = ?, net_amount = ? WHERE id = ?",
            (category_id, gst["gst_amount"], gst["net_amount"], tid),
        )
        if client_id:
            vendor_memory.remember(client_id, row["description"], category_id)
    conn.commit()
    log_audit("statement", sid, "categorize", f"{len(tids)} txns -> {cat['name']}")
    return jsonify({"updated": len(tids), "client_id": client_id})


@workflow_bp.route("/statements/<int:sid>/suggest", methods=["GET"])
def suggest(sid):
    """Vendor-memory suggestions (Part D path 1) for every uncategorized row in
    this statement. client_id resolved server-side -- no longer requires the
    frontend to pass ?client_id=, which was fragile and easy to omit."""
    conn = get_db()
    client_id = get_client_id_for_statement(conn, sid)
    rows = conn.execute(
        "SELECT * FROM transactions WHERE statement_id = ? AND category_id IS NULL", (sid,)
    ).fetchall()
    suggestions = {}
    if client_id:
        for r in rows:
            cat_id = vendor_memory.suggest_category(client_id, r["description"])
            if cat_id:
                suggestions[r["id"]] = cat_id
    return jsonify(suggestions)


# ── GST review (item 8) ─────────────────────────────────────────────────────

@workflow_bp.route("/statements/<int:sid>/gst", methods=["GET"])
def get_gst(sid):
    conn = get_db()
    rows = [_hydrate_category_fields(conn, dict(r)) for r in conn.execute(
        "SELECT * FROM transactions WHERE statement_id = ?", (sid,)
    ).fetchall()]
    return jsonify({
        "transactions": rows,
        "summary": gst_engine.summarize_gst(rows),
    })


@workflow_bp.route("/statements/<int:sid>/finalize_gst", methods=["POST"])
def finalize_gst(sid):
    conn = get_db()
    conn.execute("UPDATE statements SET status = 'gst_reviewed' WHERE id = ?", (sid,))
    conn.commit()
    log_audit("statement", sid, "finalize_gst")
    return jsonify({"status": "gst_reviewed"})


@workflow_bp.route("/statements/<int:sid>/finalize_categorize", methods=["POST"])
def finalize_categorize(sid):
    """Mark statement as categorized (step 3 complete)."""
    conn = get_db()
    conn.execute("UPDATE statements SET status = 'categorized' WHERE id = ?", (sid,))
    conn.commit()
    log_audit("statement", sid, "finalize_categorize")
    return jsonify({"status": "categorized"})


@workflow_bp.route("/statements/<int:sid>/finalize_pnl", methods=["POST"])
def finalize_pnl(sid):
    """Mark statement as fully complete (step 5 P&L done)."""
    conn = get_db()
    conn.execute("UPDATE statements SET status = 'finalized' WHERE id = ?", (sid,))
    conn.commit()
    log_audit("statement", sid, "finalize_pnl")
    return jsonify({"status": "finalized"})


# ── P&L (item 9) ─────────────────────────────────────────────────────────────

@workflow_bp.route("/statements/<int:sid>/pnl", methods=["GET"])
def get_pnl(sid):
    conn = get_db()
    rows = [_hydrate_category_fields(conn, dict(r)) for r in conn.execute(
        "SELECT * FROM transactions WHERE statement_id = ?", (sid,)
    ).fetchall()]
    return jsonify(pnl_engine.generate_pnl(rows))


@workflow_bp.route("/quarters/<int:qid>/pnl", methods=["GET"])
def get_quarter_pnl(qid):
    """Module 8: Consolidation Engine — merges all statements in a quarter."""
    conn = get_db()
    statement_ids = [r["id"] for r in conn.execute(
        "SELECT id FROM statements WHERE quarter_id = ?", (qid,)
    ).fetchall()]
    if not statement_ids:
        return jsonify({"error": "No statements in this quarter"}), 404

    placeholders = ",".join("?" * len(statement_ids))
    rows = [_hydrate_category_fields(conn, dict(r)) for r in conn.execute(
        f"SELECT * FROM transactions WHERE statement_id IN ({placeholders})", statement_ids
    ).fetchall()]
    return jsonify({
        "statement_count": len(statement_ids),
        "pnl": pnl_engine.generate_pnl(rows),
        "gst": gst_engine.summarize_gst(rows),
    })


# ── Quarters / clients (minimal, for Module 2 structure) ───────────────────

@workflow_bp.route("/business-types", methods=["GET"])
def business_types_list():
    """Powers the business_type dropdown on the client edit screen."""
    from core.business_types import list_business_types
    return jsonify(list_business_types())


@workflow_bp.route("/clients", methods=["GET", "POST"])
def clients():
    from core.business_types import is_valid_business_type
    conn = get_db()
    if request.method == "POST":
        b = request.json or {}
        business_type = b.get("business_type", "RETAIL_TRADING")
        if not is_valid_business_type(business_type):
            return jsonify({"error": f"Invalid business_type: {business_type!r}"}), 400
        cur = conn.execute(
            "INSERT INTO clients (name, business_type) VALUES (?, ?)",
            (b["name"], business_type),
        )
        conn.commit()
        log_audit("client", cur.lastrowid, "create", detail=f"business_type={business_type}")
        return jsonify({"id": cur.lastrowid})
    return jsonify([dict(r) for r in conn.execute("SELECT * FROM clients ORDER BY name").fetchall()])


@workflow_bp.route("/clients/<int:client_id>", methods=["PATCH"])
def update_client(client_id):
    """Used by the client edit screen to change name and/or business_type."""
    from core.business_types import is_valid_business_type
    conn = get_db()
    b = request.json or {}

    sets, vals = [], []
    if "name" in b:
        sets.append("name = ?")
        vals.append(b["name"])
    if "business_type" in b:
        if not is_valid_business_type(b["business_type"]):
            return jsonify({"error": f"Invalid business_type: {b['business_type']!r}"}), 400
        sets.append("business_type = ?")
        vals.append(b["business_type"])

    if not sets:
        return jsonify({"error": "No valid fields to update"}), 400

    vals.append(client_id)
    conn.execute(f"UPDATE clients SET {', '.join(sets)} WHERE id = ?", vals)
    conn.commit()
    log_audit("client", client_id, "edit", detail=str(b))

    row = conn.execute("SELECT * FROM clients WHERE id = ?", (client_id,)).fetchone()
    if row is None:
        return jsonify({"error": "Client not found"}), 404
    return jsonify(dict(row))


@workflow_bp.route("/quarters", methods=["GET", "POST"])
def quarters():
    conn = get_db()
    if request.method == "POST":
        b = request.json or {}
        cur = conn.execute(
            "INSERT INTO quarters (client_id, label, period_start, period_end) VALUES (?,?,?,?)",
            (b["client_id"], b["label"], b.get("period_start"), b.get("period_end")),
        )
        conn.commit()
        return jsonify({"id": cur.lastrowid})
    client_id = request.args.get("client_id", type=int)
    q = "SELECT * FROM quarters"
    params = ()
    if client_id:
        q += " WHERE client_id = ?"
        params = (client_id,)
    return jsonify([dict(r) for r in conn.execute(q, params).fetchall()])


# ── AI Categorize Chat (separate page, scoped to one statement) ────────────

def _build_chat_system_prompt(conn, sid: int) -> str:
    """Builds the auto-injected system context: business type, the fixed
    category list, and every currently-uncategorized transaction in this
    statement, formatted as the same plain tabular style the user
    validated manually in the Groq playground (proven to outperform JSON)."""
    client_id = get_client_id_for_statement(conn, sid)
    business_type_label = "Unknown"
    if client_id:
        row = conn.execute("SELECT business_type FROM clients WHERE id = ?", (client_id,)).fetchone()
        if row:
            from core.business_types import BUSINESS_TYPES
            business_type_label = next((b["label"] for b in BUSINESS_TYPES if b["code"] == row["business_type"]), "Unknown")

    categories = category_master.list_categories()
    cat_list_str = "\n".join(f'- "{c["name"]}"' for c in categories)

    txns = conn.execute(
        "SELECT * FROM transactions WHERE statement_id = ? AND category_id IS NULL", (sid,)
    ).fetchall()
    txn_lines = "\n".join(
        f'{t["id"]}\t{t["date"] or ""}\t{t["description"]}\t{t["amount"]}\t'
        f'{"DR" if t["amount"] < 0 else "CR"}'
        for t in txns
    )

    return (
        "You are a bookkeeping categorization assistant for an Australian "
        "accounting firm, helping categorize bank transactions for GST BAS. "
        f"Business type: {business_type_label}.\n\n"
        "Stick to exactly these categories, never invent new ones, copy "
        "names exactly:\n"
        f"{cat_list_str}\n\n"
        "A 'DR' transaction is a debit (money out, an expense) -- it must "
        "map to an Expense or Excluded category, never Income. A 'CR' "
        "transaction is a credit (money in, income) -- it must map to "
        "Income or Excluded, never Expense.\n\n"
        "Uncategorized transactions in this statement "
        "(id, date, description, amount, direction):\n"
        f"{txn_lines}\n\n"
        "When the user asks you to categorize, respond with a numbered "
        "list: id - category. Keep responses focused on this task."
    )


@workflow_bp.route("/statements/<int:sid>/ai-chat", methods=["POST"])
def ai_chat(sid):
    """
    body: {messages: [{role: 'user'|'assistant', content: str}, ...]}
    The system prompt (business context + fixed categories + uncategorized
    transactions) is injected server-side automatically -- the frontend
    only ever sends the user-visible conversation turns.

    Returns the AI's reply plus REAL usage/rate-limit data from Groq for
    the analytics panel. This endpoint only returns text -- it never
    writes to the database. Applying any proposed categorization still
    goes through the existing /categorize endpoint, same as every other
    path in this app.
    """
    conn = get_db()
    body = request.json or {}
    chat_messages = body.get("messages", [])

    system_prompt = _build_chat_system_prompt(conn, sid)
    full_messages = [{"role": "system", "content": system_prompt}] + chat_messages

    result = category_engine.call_groq_with_usage(full_messages, max_tokens=1500)

    usage = result.get("usage") or {}
    rl = result.get("rate_limit") or {}
    conn.execute(
        """INSERT INTO ai_usage_log
           (statement_id, prompt_tokens, completion_tokens, total_tokens,
            limit_requests, remaining_requests, limit_tokens, remaining_tokens)
           VALUES (?,?,?,?,?,?,?,?)""",
        (sid, usage.get("prompt_tokens"), usage.get("completion_tokens"), usage.get("total_tokens"),
         rl.get("limit_requests"), rl.get("remaining_requests"), rl.get("limit_tokens"), rl.get("remaining_tokens")),
    )
    conn.commit()

    if result["error"]:
        return jsonify({"error": result["error"]}), 502

    return jsonify({"reply": result["content"], "usage": usage, "rate_limit": rl})


@workflow_bp.route("/ai-usage/summary", methods=["GET"])
def ai_usage_summary():
    """Powers the analytics panel under the AI chat page. All numbers here
    are either real totals from our own log table, or the most recent
    real rate-limit snapshot Groq sent us -- nothing estimated."""
    conn = get_db()
    today_row = conn.execute(
        """SELECT COUNT(*) requests_today, COALESCE(SUM(total_tokens),0) tokens_today
           FROM ai_usage_log WHERE date(created_at) = date('now')"""
    ).fetchone()
    all_time_row = conn.execute(
        """SELECT COUNT(*) requests_all_time, COALESCE(SUM(total_tokens),0) tokens_all_time
           FROM ai_usage_log"""
    ).fetchone()
    latest = conn.execute(
        """SELECT limit_requests, remaining_requests, limit_tokens, remaining_tokens, created_at
           FROM ai_usage_log WHERE limit_requests IS NOT NULL
           ORDER BY id DESC LIMIT 1"""
    ).fetchone()

    return jsonify({
        "requests_today": today_row["requests_today"],
        "tokens_today": today_row["tokens_today"],
        "requests_all_time": all_time_row["requests_all_time"],
        "tokens_all_time": all_time_row["tokens_all_time"],
        "latest_rate_limit": dict(latest) if latest else None,
        "note": "Groq does not expose your daily request cap (RPD) in response "
                "headers -- 'requests_today' above is tracked locally from our "
                "own call log, not from Groq. Per-minute remaining requests/tokens "
                "in 'latest_rate_limit' ARE real values from Groq's last response.",
    })


# ── Raw AI Playground (zero injected context — pure passthrough to Groq) ───

# ── AI Categorize Page — prompt generation + paste-back apply ────────────────

# ─── Australian vendor hints ────────────────────────────────────────────────
# Injected once into every prompt — ~200 tokens, dramatically reduces misclassification
# of well-known AU brands.  Sea World is a theme park, Telstra is telco, etc.
_AU_VENDOR_HINTS = """
Australian vendor reference (strong hints — use these when description matches):
TRAVEL & TRANSPORT: Uber, Ola, DiDi, 13Cabs, Silver Top, Cabcharge, Qantas, Virgin Australia, Jetstar, Rex Airlines, Hertz, Avis, Budget Car Rental, Europcar, Thrifty, GoGet, Translink, Opal, Myki, Metro Trains, Sydney Trains, Yarra Trams, Transperth, Adelaide Metro → Travel & Vehicle
THEME PARKS & ENTERTAINMENT (leisure, NOT food): Sea World, Movie World, Wet'n'Wild, Dreamworld, WhiteWater World, Luna Park, Taronga Zoo, Melbourne Zoo, Wildlife Sydney, SEALIFE Aquarium, Madame Tussauds, Timezone, Strike Bowling, Event Cinemas, Hoyts, Village Cinemas, Palace Cinemas → Travel & Vehicle (or Office & Operating Expenses if a business event)
GROCERIES (Food & Meals): Woolworths, Coles, IGA, ALDI, Costco, Harris Farm, Foodland, Drakes, Ritchies → Food & Meals
OFFICE & STATIONERY: Officeworks, Staples → Office & Operating Expenses
TELCO: Telstra, Optus, Vodafone, TPG, iiNet, Aussie Broadband, Belong → Office & Operating Expenses
UTILITIES: AGL, Origin Energy, EnergyAustralia, Red Energy, Alinta Energy, Ergon, Ausgrid → Office & Operating Expenses
INSURANCE: NRMA, AAMI, Allianz, QBE, CGU, Budget Direct, GIO, RAA, RAC, RACQ, RACV, Medibank, Bupa, HCF, NIB, HBF → Insurance
BANK FEES: any "account fee", "overdrawn fee", "dishonour fee", "monthly fee", "card fee", "bpay fee" — only if it is a DR (debit). A CR at a bank is income or a transfer, never a fee.
PROFESSIONAL: MYOB, Xero, QuickBooks, Reckon, law firms, barristers, consultants → Professional / Contractor Fees
""".strip()

import re as _re

def _clean_description(desc: str) -> str:
    """
    Remove NAB-style dot-padding (e.g. 'Sea World Resort.......... 110.00')
    and collapse multiple spaces. Bank statement formatting noise only.
    """
    if not desc:
        return ""
    # Remove trailing dot sequences (3+ dots) and everything after them on the same token
    cleaned = _re.sub(r'\.{3,}.*$', '', desc)
    # Collapse whitespace
    cleaned = _re.sub(r'\s+', ' ', cleaned).strip()
    return cleaned


BATCH_SIZE = 150   # max transactions per prompt batch

@workflow_bp.route("/statements/<int:sid>/ai-categorize/prompt", methods=["GET"])
def ai_categorize_prompt(sid):
    """
    Returns prompt data for the AI Categorize page.
    If > BATCH_SIZE uncategorized transactions exist, returns multiple prompt batches.
    Each batch uses:
      - Real transaction_id labels (e.g. nab_0001)
      - Cleaned descriptions (NAB dot-padding stripped)
      - Date-ordered rows
      - Australian vendor hints section
      - Strict direction rules per category
    No AI call is made here.
    """
    conn = get_db()
    client_id = get_client_id_for_statement(conn, sid)
    business_type_label = "Unknown"
    if client_id:
        row = conn.execute("SELECT business_type FROM clients WHERE id = ?", (client_id,)).fetchone()
        if row:
            from core.business_types import BUSINESS_TYPES
            business_type_label = next(
                (b["label"] for b in BUSINESS_TYPES if b["code"] == row["business_type"]), "Unknown"
            )

    categories = category_master.list_categories()
    cat_names = [c["name"] for c in categories]

    # Split categories by direction for the strict rules section
    income_cats      = [c["name"] for c in categories if c.get("pnl_group") == "Income"]
    expense_cats     = [c["name"] for c in categories if c.get("pnl_group") == "Expense"]
    direct_cost_cats = [c["name"] for c in categories if c.get("pnl_group") == "Direct Cost"]
    excl_cats        = [c["name"] for c in categories if c.get("pnl_group") == "Excluded"]

    cat_list_str = "\n".join(f"  - {n}" for n in cat_names)

    # Order by date ASC, then by DB id to break ties — ensures chronological sequence
    txns = conn.execute(
        """SELECT * FROM transactions
           WHERE statement_id = ? AND category_id IS NULL
           ORDER BY date ASC, id ASC""",
        (sid,)
    ).fetchall()

    if not txns:
        return jsonify({
            "batches": [],
            "categories": cat_names,
            "total_uncategorized": 0,
            "message": "No uncategorized transactions — nothing to do."
        })

    # Build clean row list once
    all_rows = []
    for t in txns:
        tid_label = t["transaction_id"] or str(t["id"])
        direction = "DR" if t["amount"] < 0 else "CR"
        clean_desc = _clean_description(t["description"])
        all_rows.append({
            "tid": tid_label,
            "date": t["date"] or "",
            "desc": clean_desc,
            "amount": abs(t["amount"]),
            "dir": direction,
        })

    # Split into batches of BATCH_SIZE
    batches = [all_rows[i:i+BATCH_SIZE] for i in range(0, len(all_rows), BATCH_SIZE)]
    total_batches = len(batches)

    def build_prompt(batch, batch_num, total):
        header = "ID\tDate\tDescription\tAmount\tDR/CR"
        rows_text = "\n".join(
            f'{r["tid"]}\t{r["date"]}\t{r["desc"]}\t{r["amount"]:.2f}\t{r["dir"]}'
            for r in batch
        )
        batch_label = f" (Batch {batch_num}/{total})" if total > 1 else ""
        example_ids = [r["tid"] for r in batch[:3]]
        example_cats = ["Travel & Vehicle", "Salary & Wages", "Sales / Trading Income"]
        example_lines = "\n".join(
            f"{eid}: {ecat}" for eid, ecat in zip(example_ids, example_cats)
        )
        return (
            f"You are an Australian bookkeeping assistant{batch_label}.\n"
            f"Business type: {business_type_label}\n\n"
            f"=== CATEGORIES (use EXACT names, case-sensitive) ===\n"
            f"{cat_list_str}\n\n"
            f"=== DIRECTION RULES (HARD — never break these) ===\n"
            f"DR (debit, money OUT) → ONLY these groups: Direct Cost, Expense, or Excluded\n"
            f"  Direct Cost: {', '.join(direct_cost_cats)}\n"
            f"  Expense: {', '.join(expense_cats)}\n"
            f"  Excluded: {', '.join(excl_cats)}\n"
            f"CR (credit, money IN) → ONLY these groups: Income or Excluded\n"
            f"  Income: {', '.join(income_cats)}\n"
            f"  Excluded: {', '.join(excl_cats)}\n"
            f"NEVER assign an Income or Direct Cost category to a CR row. NEVER assign an Expense or Direct Cost category to a CR row.\n\n"
            f"=== AUSTRALIAN VENDOR HINTS ===\n"
            f"{_AU_VENDOR_HINTS}\n\n"
            f"=== TRANSACTIONS{batch_label} ===\n"
            f"{header}\n"
            f"{rows_text}\n\n"
            f"=== YOUR RESPONSE ===\n"
            f"Output ONLY one line per transaction in this exact format, nothing else:\n"
            f"{example_lines}"
        )

    prompts = [
        {
            "batch_num": i + 1,
            "total_batches": total_batches,
            "count": len(batch),
            "label": f"Batch {i+1} of {total_batches} ({len(batch)} transactions)" if total_batches > 1 else f"{len(batch)} transactions",
            "prompt": build_prompt(batch, i + 1, total_batches),
        }
        for i, batch in enumerate(batches)
    ]

    return jsonify({
        "batches": prompts,
        "categories": cat_names,
        "total_uncategorized": len(txns),
        "business_type": business_type_label
    })


@workflow_bp.route("/statements/<int:sid>/ai-categorize/apply", methods=["POST"])
def ai_categorize_apply(sid):
    """
    body: {response_text: "101: Travel & Transport\\n102: Salary & Wages\\n..."}

    Parses the pasted AI response (plain id: category_name format),
    validates each line against the category master, applies the direction
    guardrail, writes to DB via the same GST recalc path as /categorize,
    and returns a detailed result summary.

    One malformed line never breaks the rest — parsing is line-by-line.
    """
    import re
    conn = get_db()
    body = request.json or {}
    response_text = body.get("response_text", "").strip()

    if not response_text:
        return jsonify({"error": "No response text provided"}), 400

    categories = category_master.list_categories()
    # Build name -> category lookup (case-insensitive)
    cat_by_name = {c["name"].lower().strip(): c for c in categories}

    # Build transaction_id -> db row lookup (primary), and db id -> row (fallback)
    txns = conn.execute(
        "SELECT * FROM transactions WHERE statement_id = ?", (sid,)
    ).fetchall()
    txn_by_txnid = {}   # e.g. "nab001" -> row
    txn_by_dbid  = {}   # e.g. 1870 -> row (fallback)
    for t in txns:
        if t["transaction_id"]:
            txn_by_txnid[t["transaction_id"]] = t
        txn_by_dbid[t["id"]] = t

    client_id = get_client_id_for_statement(conn, sid)

    applied = []
    skipped = []
    errors = []

    for line in response_text.splitlines():
        line = line.strip()
        if not line:
            continue

        # Match "nab001: Category Name" or "1870: Category Name" or "nab001 - Category Name"
        # ID can be alphanumeric (e.g. nab001, cba002) or purely numeric
        m = re.match(r'^([A-Za-z0-9_\-]+)\s*[:\-]\s*(.+)$', line)
        if not m:
            errors.append({"line": line, "reason": "Could not parse format (expected: ID: Category Name)"})
            continue

        raw_id = m.group(1).strip()
        cat_name = m.group(2).strip()

        # Look up transaction — prefer transaction_id, fall back to db id
        txn = txn_by_txnid.get(raw_id)
        if not txn:
            # Try numeric db id fallback
            try:
                txn = txn_by_dbid.get(int(raw_id))
            except ValueError:
                pass
        if not txn:
            errors.append({"line": line, "reason": f'ID "{raw_id}" not found in this statement'})
            continue

        # Use DB row id for updates
        db_tid = txn["id"]

        # Look up category (case-insensitive)
        cat = cat_by_name.get(cat_name.lower())
        if not cat:
            # Try fuzzy: check if the AI name is a substring of a real name
            matches = [c for name, c in cat_by_name.items() if cat_name.lower() in name or name in cat_name.lower()]
            if len(matches) == 1:
                cat = matches[0]
            else:
                errors.append({"line": line, "reason": f'Category "{cat_name}" not recognized'})
                continue

        # Direction guardrail — same logic as existing categorize endpoint
        amount = txn["amount"]
        is_debit = amount < 0
        pnl_group = cat.get("pnl_group", "")
        direction_ok = True
        if is_debit and pnl_group == "Income":
            direction_ok = False
        elif not is_debit and pnl_group == "Expense":
            direction_ok = False

        if not direction_ok:
            direction = "DR" if is_debit else "CR"
            errors.append({
                "line": line,
                "reason": f"Direction mismatch: {direction} transaction cannot be {pnl_group} category"
            })
            continue

        # Apply — same GST recalc path as /categorize
        gst = gst_engine.recalc_transaction_gst(amount, cat)
        conn.execute(
            "UPDATE transactions SET category_id = ?, gst_amount = ?, net_amount = ? WHERE id = ?",
            (cat["id"], gst["gst_amount"], gst["net_amount"], db_tid),
        )
        if client_id:
            vendor_memory.remember(client_id, txn["description"], cat["id"])

        applied.append({"id": raw_id, "description": txn["description"], "category": cat["name"]})

    conn.commit()

    if applied:
        log_audit("statement", sid, "ai-categorize-apply",
                  f"{len(applied)} applied, {len(skipped)} skipped, {len(errors)} errors")

    return jsonify({
        "applied": len(applied),
        "applied_detail": applied,
        "errors": errors,
        "skipped": skipped,
        "total_lines": len([l for l in response_text.splitlines() if l.strip()])
    })


@workflow_bp.route("/ai-playground/chat", methods=["POST"])
def ai_playground_chat():
    """
    body: {messages: [{role: 'user'|'assistant', content: str}, ...]}
    NO system prompt, NO transaction data, NO category list, NO business
    context is injected here -- whatever the frontend sends is exactly
    what goes to Groq, nothing more. This exists purely so usage/token
    behavior can be compared directly against the Groq playground with
    no project-specific scaffolding in the way.
    """
    body = request.json or {}
    chat_messages = body.get("messages", [])
    result = category_engine.call_groq_with_usage(chat_messages, max_tokens=1500)

    usage = result.get("usage") or {}
    rl = result.get("rate_limit") or {}
    conn = get_db()
    conn.execute(
        """INSERT INTO ai_usage_log
           (statement_id, prompt_tokens, completion_tokens, total_tokens,
            limit_requests, remaining_requests, limit_tokens, remaining_tokens)
           VALUES (NULL,?,?,?,?,?,?,?)""",
        (usage.get("prompt_tokens"), usage.get("completion_tokens"), usage.get("total_tokens"),
         rl.get("limit_requests"), rl.get("remaining_requests"), rl.get("limit_tokens"), rl.get("remaining_tokens")),
    )
    conn.commit()

    if result["error"]:
        return jsonify({"error": result["error"]}), 502
    return jsonify({"reply": result["content"], "usage": usage, "rate_limit": rl})


# ── Quarters/Statements list (new) ─────────────────────────────────────────

@workflow_bp.route("/quarters/<int:qid>/statements", methods=["GET"])
def get_quarter_statements(qid):
    conn = get_db()
    rows = conn.execute("""
        SELECT s.*, COUNT(t.id) as txn_count
        FROM statements s
        LEFT JOIN transactions t ON t.statement_id = s.id
        WHERE s.quarter_id = ?
        GROUP BY s.id
        ORDER BY s.created_at DESC
    """, (qid,)).fetchall()
    return jsonify([dict(r) for r in rows])


@workflow_bp.route("/statements/<int:sid>", methods=["DELETE"])
def delete_statement(sid):
    conn = get_db()
    conn.execute("DELETE FROM transactions WHERE statement_id = ?", (sid,))
    conn.execute("DELETE FROM statements WHERE id = ?", (sid,))
    conn.commit()
    return jsonify({"deleted": sid})


# ── Statement name update ───────────────────────────────────────────────────

@workflow_bp.route("/statements/<int:sid>/name", methods=["PATCH"])
def update_statement_name(sid):
    b = request.json or {}
    name = b.get("statement_name", "").strip()
    conn = get_db()
    conn.execute("UPDATE statements SET statement_name = ? WHERE id = ?", (name, sid))
    conn.commit()
    return jsonify({"id": sid, "statement_name": name})


# ── Quarter consolidation ───────────────────────────────────────────────────

@workflow_bp.route("/quarters/<int:qid>/consolidate", methods=["POST"])
def consolidate_quarter(qid):
    """Merge transactions from multiple statements into one consolidated statement."""
    b = request.json or {}
    stmt_ids = b.get("statement_ids", [])
    name = b.get("name", "Consolidated BAS Report")
    if not stmt_ids:
        return jsonify({"error": "No statement IDs provided"}), 400

    conn = get_db()

    # Create a new consolidated statement
    cur = conn.execute(
        "INSERT INTO statements (quarter_id, bank_id, filename, statement_name, status) VALUES (?, 'consolidated', ?, ?, 'parsed')",
        (qid, name, name)
    )
    new_stmt_id = cur.lastrowid

    # Copy all transactions from selected statements
    placeholders = ",".join("?" * len(stmt_ids))
    txns = conn.execute(
        f"SELECT * FROM transactions WHERE statement_id IN ({placeholders}) ORDER BY date, id",
        stmt_ids
    ).fetchall()

    count = 0
    for t in txns:
        conn.execute("""
            INSERT INTO transactions (statement_id, transaction_id, date, description,
                amount, balance, source_page, row_top, confidence, approved,
                category_id, gst_amount, net_amount, group_key)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            new_stmt_id,
            f"consol_{new_stmt_id}_{t['id']}",
            t["date"], t["description"], t["amount"], t["balance"],
            t["source_page"], t["row_top"], t["confidence"], t["approved"],
            t["category_id"], t["gst_amount"], t["net_amount"], t["group_key"]
        ))
        count += 1

    conn.commit()
    log_audit("statement", new_stmt_id, "consolidate", f"Merged {len(stmt_ids)} statements, {count} transactions")
    return jsonify({"consolidated_statement_id": new_stmt_id, "txn_count": count, "name": name})


# ── Annual consolidation ────────────────────────────────────────────────────

@workflow_bp.route("/consolidate/annual", methods=["POST"])
def consolidate_annual():
    import json
    b = request.json or {}
    client_id = b.get("client_id")
    label = b.get("label", "Annual")
    quarter_ids = b.get("quarter_ids", [])
    if not quarter_ids or not client_id:
        return jsonify({"error": "client_id and quarter_ids required"}), 400

    conn = get_db()

    # Collect all transactions across all quarters
    all_count = 0
    for qid in quarter_ids:
        stmts = conn.execute("SELECT id FROM statements WHERE quarter_id = ?", (qid,)).fetchall()
        for s in stmts:
            cnt = conn.execute("SELECT COUNT(*) FROM transactions WHERE statement_id = ?", (s["id"],)).fetchone()[0]
            all_count += cnt

    # Store the annual consolidation record
    cur = conn.execute(
        "INSERT INTO annual_consolidations (client_id, label, quarter_ids) VALUES (?, ?, ?)",
        (client_id, label, json.dumps(quarter_ids))
    )
    conn.commit()

    return jsonify({"annual_id": cur.lastrowid, "label": label, "txn_count": all_count, "quarters": len(quarter_ids)})


# ── Business types (for dropdown) ──────────────────────────────────────────

@workflow_bp.route("/business-types", methods=["GET"])
def business_types():
    try:
        from core.business_types import BUSINESS_TYPES
        return jsonify([{"code": k, "label": v} for k, v in BUSINESS_TYPES.items()])
    except Exception:
        return jsonify([
            {"code": "RETAIL_TRADING", "label": "Retail Trading"},
            {"code": "SERVICE", "label": "Service Business"},
            {"code": "PROFESSIONAL", "label": "Professional Services"},
            {"code": "CONSTRUCTION", "label": "Construction"},
            {"code": "HOSPITALITY", "label": "Hospitality"},
        ])



@workflow_bp.route("/import/headers", methods=["POST"])
def get_import_headers():
    """Read uploaded CSV/Excel and return column headers + sample row for mapping UI."""
    import io as _io
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400
    f    = request.files["file"]
    raw  = f.read()
    fname = (f.filename or "").lower()
    rows = []
    try:
        if fname.endswith(".csv"):
            import csv
            for enc in ("utf-8-sig", "utf-8", "latin-1"):
                try:
                    text = raw.decode(enc)
                    rows = [dict(r) for r in csv.DictReader(_io.StringIO(text))]
                    break
                except Exception:
                    continue
        elif fname.endswith((".xlsx", ".xls")):
            import openpyxl
            wb  = openpyxl.load_workbook(_io.BytesIO(raw), data_only=True)
            ws  = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if all_rows:
                headers = [str(c).strip() if c is not None else f"col_{i}"
                           for i, c in enumerate(all_rows[0])]
                rows = [
                    {h: (str(v).strip() if v is not None else "")
                     for h, v in zip(headers, row)}
                    for row in all_rows[1:]
                ]
        else:
            return jsonify({"error": "Unsupported file type. Use .csv, .xlsx or .xls"}), 400
    except Exception as e:
        return jsonify({"error": f"Could not read file: {e}"}), 400

    if not rows:
        return jsonify({"error": "File appears empty"}), 400

    headers = list(rows[0].keys())
    return jsonify({"headers": headers, "sample": rows[:3], "row_count": len(rows)})


@workflow_bp.route("/statements/import-csv", methods=["POST"])
def import_csv_excel():
    """
    Import transactions from a CSV or Excel file.

    Accepts .csv, .xlsx, .xls — any column header naming, any case.

    Date column:   date, transaction date, txn date, value date, posted date, trans date
    Desc column:   description, details, particulars, narrative, memo, reference,
                   transaction details, trans desc, narration, remarks, note
    Amount cols:   debit+credit (separate), or amount, debit/credit, net amount,
                   withdrawal+deposit, withdrawals+deposits
    """
    import io as _io
    from datetime import datetime as _dt

    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    f          = request.files["file"]
    fname      = (f.filename or "").lower()
    name       = request.form.get("name", "").strip() or (f.filename or "Imported").rsplit(".", 1)[0]
    client_id  = request.form.get("client_id",  type=int)
    quarter_id = request.form.get("quarter_id", type=int)

    raw = f.read()

    # ── Parse file into list of dicts ─────────────────────────────────────────
    try:
        if fname.endswith(".csv"):
            import csv
            # Try UTF-8 with BOM first, fall back to latin-1
            for enc in ("utf-8-sig", "utf-8", "latin-1"):
                try:
                    text = raw.decode(enc, errors="strict")
                    break
                except (UnicodeDecodeError, LookupError):
                    continue
            else:
                text = raw.decode("latin-1", errors="replace")
            reader = list(csv.DictReader(_io.StringIO(text)))
            rows   = [dict(r) for r in reader]

        elif fname.endswith((".xlsx", ".xls")):
            try:
                import openpyxl
                wb  = openpyxl.load_workbook(_io.BytesIO(raw), data_only=True)
                ws  = wb.active
                all_rows = list(ws.iter_rows(values_only=True))
                if not all_rows:
                    return jsonify({"error": "Empty spreadsheet"}), 400
                headers = [str(c).strip() if c is not None else f"col_{i}"
                           for i, c in enumerate(all_rows[0])]
                rows = [
                    {h: (str(v).strip() if v is not None else "")
                     for h, v in zip(headers, row)}
                    for row in all_rows[1:]
                ]
            except ImportError:
                return jsonify({"error": "openpyxl not installed — cannot read Excel files. "
                                         "Run: pip install openpyxl"}), 500
        else:
            return jsonify({"error": "Unsupported file type. Use .csv, .xlsx or .xls"}), 400

    except Exception as e:
        return jsonify({"error": f"Could not read file: {e}"}), 400

    if not rows:
        return jsonify({"error": "File has no data rows"}), 400

    # ── Column detection — use explicit mapping if provided, else auto-detect ──
    import json as _json
    mapping_raw = request.form.get("mapping")
    mapping     = None
    if mapping_raw:
        try:
            mapping = _json.loads(mapping_raw)
        except Exception:
            pass

    def _find_col(headers, *candidates):
        normalise = lambda s: " ".join(s.lower().split())
        norm_map  = {normalise(h): h for h in headers}
        for c in candidates:
            hit = norm_map.get(normalise(c))
            if hit is not None:
                return hit
        for c in candidates:
            nc = normalise(c)
            for nh, orig in norm_map.items():
                if nc in nh or nh in nc:
                    return orig
        return None

    headers = list(rows[0].keys())

    if mapping:
        # User provided explicit mapping from the UI
        col_date  = mapping.get("date")  or None
        col_desc  = mapping.get("description") or None
        col_amt   = mapping.get("amount") or None
        col_deb   = mapping.get("debit")  or None
        col_cred  = mapping.get("credit") or None
    else:
        # Auto-detect
        col_date  = _find_col(headers,
                        "date", "transaction date", "txn date", "value date",
                        "posted date", "trans date", "settlement date", "booking date")
        col_desc  = _find_col(headers,
                        "description", "details", "particulars", "narrative",
                        "memo", "reference", "transaction details", "trans desc",
                        "narration", "remarks", "note", "notes", "transaction description",
                        "payment details", "transaction narrative", "payee")
        col_deb   = _find_col(headers,
                        "debit", "withdrawal", "withdrawals", "debit amount",
                        "dr", "debit (aud)", "amount dr")
        col_cred  = _find_col(headers,
                        "credit", "deposit", "deposits", "credit amount",
                        "cr", "credit (aud)", "amount cr")
        col_amt   = _find_col(headers,
                        "amount", "debit/credit", "credit/debit", "net amount",
                        "transaction amount", "value", "net", "dr/cr amount")

    # Validation
    if not col_date:
        found = ", ".join(f"'{h}'" for h in headers[:8])
        return jsonify({"error": f"Cannot find a Date column. Columns found: {found}"}), 400
    if not col_desc:
        found = ", ".join(f"'{h}'" for h in headers[:8])
        return jsonify({"error": f"Cannot find a Description column. Columns found: {found}"}), 400
    if not col_deb and not col_cred and not col_amt:
        found = ", ".join(f"'{h}'" for h in headers[:8])
        return jsonify({"error": f"Cannot find an Amount column. Columns found: {found}"}), 400

    # ── Parse money ───────────────────────────────────────────────────────────
    def _money(s):
        if s is None: return None
        s = str(s).strip().replace(",", "").replace("$", "").replace(" ", "")
        if not s or s in ("-", "—", "–", ""): return None
        # Handle (1234.56) accounting negatives
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        try: return float(s)
        except ValueError: return None

    # ── Parse date ────────────────────────────────────────────────────────────
    def _date(s):
        if not s: return None
        s = str(s).strip()
        # Excel serial number (float/int stored as string)
        try:
            serial = float(s)
            if 1000 < serial < 100000:
                from datetime import date as _date_cls, timedelta
                # Excel epoch: 1899-12-30
                return (_date_cls(1899, 12, 30) + timedelta(days=int(serial))).strftime("%d-%b-%Y")
        except ValueError:
            pass
        for fmt in (
            "%d/%m/%Y", "%d/%m/%y",
            "%Y-%m-%d", "%Y/%m/%d",
            "%d-%m-%Y", "%d-%m-%y",
            "%d %b %Y", "%d %b %y",
            "%d-%b-%Y", "%d-%b-%y",
            "%d %B %Y", "%d %B %y",
            "%m/%d/%Y", "%m/%d/%y",
            "%d.%m.%Y", "%d.%m.%y",
            "%b %d, %Y", "%B %d, %Y",
            "%Y%m%d",
        ):
            try: return _dt.strptime(s, fmt).strftime("%d-%b-%Y")
            except ValueError: continue
        return s  # keep raw if unparseable

    # ── Build transaction list ────────────────────────────────────────────────
    transactions = []
    skipped      = 0

    for i, row in enumerate(rows):
        date_raw = str(row.get(col_date, "") or "").strip()
        desc     = str(row.get(col_desc, "") or "").strip()

        # Skip completely empty rows
        if not date_raw and not desc:
            skipped += 1
            continue

        parsed_date = _date(date_raw)

        if col_deb and col_cred:
            dv = _money(row.get(col_deb))
            cv = _money(row.get(col_cred))
            amount = round((-abs(dv) if dv else 0.0) + (abs(cv) if cv else 0.0), 2)
        elif col_amt:
            amount = round(_money(row.get(col_amt)) or 0.0, 2)
        else:
            # Fallback: try debit only or credit only
            dv = _money(row.get(col_deb)) if col_deb else None
            cv = _money(row.get(col_cred)) if col_cred else None
            amount = round((-abs(dv) if dv else 0.0) + (abs(cv) if cv else 0.0), 2)

        if amount == 0.0 and not desc:
            skipped += 1
            continue

        transactions.append({
            "transaction_id": "",
            "date":        parsed_date or "",
            "description": desc,
            "amount":      amount,
            "balance":     None,
            "source_page": 1,
            "row_top":     float(i),
            "confidence":  1.0,
        })

    if not transactions:
        return jsonify({"error": f"No valid transactions found ({skipped} rows skipped). "
                                 "Check the file has data rows and matching column names."}), 400

    # Sort by date, then assign IDs
    try:
        transactions.sort(
            key=lambda t: (_dt.strptime(t["date"], "%d-%b-%Y") if t["date"] else _dt.min,
                           t["row_top"])
        )
    except Exception:
        pass

    for i, t in enumerate(transactions):
        t["transaction_id"] = f"import_{i+1:04d}"

    # ── Persist to DB ─────────────────────────────────────────────────────────
    conn = get_db()

    quarter_id = request.form.get("quarter_id") or None
    name       = request.form.get("name") or (f.filename or "import").rsplit(".", 1)[0]

    stmt_id = conn.execute(
        "INSERT INTO statements (quarter_id, statement_name, bank_id, filename, status, created_at) "
        "VALUES (?,?,?,?,?,?)",
        (quarter_id, name, "import", f.filename or name, "parsed",
         __import__("datetime").datetime.utcnow().isoformat())
    ).lastrowid
    conn.commit()

    for t in transactions:
        conn.execute(
            """INSERT INTO transactions
               (statement_id, transaction_id, date, description, amount, balance,
                source_page, row_top, confidence)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (stmt_id, t["transaction_id"], t["date"], t["description"],
             t["amount"], t["balance"], t["source_page"], t["row_top"], t["confidence"])
        )
    conn.commit()

    saved = [dict(r) for r in conn.execute(
        "SELECT * FROM transactions WHERE statement_id = ? ORDER BY id", (stmt_id,)
    ).fetchall()]

    return jsonify({
        "statement_id": stmt_id,
        "transactions": saved,
        "count":        len(saved),
        "skipped":      skipped,
        "name":         name,
    })
