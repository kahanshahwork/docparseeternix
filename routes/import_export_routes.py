"""
import_export_routes.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Handles:
  POST /api/import/csv              — import CSV or Excel → statement + transactions
  POST /api/statements/import-csv  — alias (same handler, JS-friendly URL)
  GET  /api/statements/<id>/export/gst  — download GST summary as .xlsx
  GET  /api/statements/<id>/export/pnl  — download P&L as .xlsx
"""

import io
import os
import re
import csv
from datetime import datetime

from flask import Blueprint, jsonify, request, send_file
from core.db import get_db, log_audit
from core import gst_engine, pnl_engine

ie_bp = Blueprint("import_export", __name__, url_prefix="/api")


# ── helpers ──────────────────────────────────────────────────────────────────

def _parse_money(s):
    """Convert '$-1,650.00' / '-1650' / '' / float → float or None."""
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    s = str(s).strip()
    if not s or s in ("-", "—", "N/A", "n/a"):
        return None
    s = re.sub(r"[^\d.\-]", "", s)
    try:
        return float(s)
    except ValueError:
        return None


# Flexible header aliases — covers many real-world bank export formats
_DATE_KEYS = [
    "Date", "date", "DATE",
    "Trans Date", "Transaction Date", "Trans. Date", "Txn Date",
    "Value Date", "Settlement Date", "Posting Date",
]
_DESC_KEYS = [
    "Description", "description", "DESCRIPTION",
    "Descriptions",
    "Narrative", "narrative", "NARRATIVE",
    "Narration", "narration", "NARRATION",
    "Details", "details", "DETAILS",
    "Transaction Details", "Transaction details", "Transaction Description",
    "Particulars", "particulars", "PARTICULARS",
    "Memo", "memo", "MEMO",
    "Reference", "reference", "REFERENCE",
    "Remarks", "remarks",
    "Transaction Remarks", "Transaction Reference",
    "Info", "Additional Info",
    "Payee", "Payee/Description",
]
_DEBIT_KEYS = [
    "Debit", "debit", "DEBIT",
    "Withdrawals", "Withdrawal", "DR", "Dr",
    "Money Out", "Debit Amount", "Paid Out",
    "Cheques", "Cheque",
]
_CREDIT_KEYS = [
    "Credit", "credit", "CREDIT",
    "Deposits", "Deposit", "CR", "Cr",
    "Money In", "Credit Amount", "Paid In",
    "Receipts",
]
_AMOUNT_KEYS = [
    "Amount", "amount", "AMOUNT",
    "Net Amount", "Transaction Amount", "Value",
]


def _get_field(raw, keys):
    """Try a list of key aliases against a row dict, return first match.
    Always does case-insensitive match so Transaction details == Transaction Details."""
    # Build lower-keyed version of row once
    raw_lower = {k2.lower().strip(): v2 for k2, v2 in raw.items()}
    # First pass: exact match
    for k in keys:
        v = raw.get(k)
        if v is not None and str(v).strip() not in ("", "None"):
            return str(v).strip()
    # Second pass: case-insensitive
    for k in keys:
        v = raw_lower.get(k.lower().strip())
        if v is not None and str(v).strip() not in ("", "None"):
            return str(v).strip()
    # Third pass: partial match (e.g. "transaction" matches "Transaction details")
    for raw_key, raw_val in raw.items():
        if raw_val is None or str(raw_val).strip() in ("", "None"):
            continue
        rk = raw_key.lower().strip()
        for k in keys:
            if k.lower().strip() in rk or rk in k.lower().strip():
                return str(raw_val).strip()
    return ""


def _normalise_row(raw, idx):
    """Convert a dict from CSV/Excel → canonical transaction dict."""
    date        = _get_field(raw, _DATE_KEYS)
    description = _get_field(raw, _DESC_KEYS)
    debit_raw   = _get_field(raw, _DEBIT_KEYS)
    credit_raw  = _get_field(raw, _CREDIT_KEYS)
    amount_raw  = _get_field(raw, _AMOUNT_KEYS)

    if debit_raw or credit_raw:
        debit_val  = _parse_money(debit_raw)
        credit_val = _parse_money(credit_raw)
        if credit_val and credit_val != 0:
            amount = abs(credit_val)
        elif debit_val and debit_val != 0:
            amount = -abs(debit_val)
        else:
            amount = 0.0
    elif amount_raw:
        amount = _parse_money(amount_raw) or 0.0
    else:
        amount = 0.0

    return {
        "date":        date,
        "description": description,
        "amount":      amount,
        "source_page": None,
    }


def _read_csv(file_bytes, filename):
    """Parse CSV bytes → list of row dicts."""
    # Try UTF-8-sig first (Excel CSV), then latin-1 fallback
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = file_bytes.decode(enc, errors="strict")
            break
        except (UnicodeDecodeError, LookupError):
            continue
    else:
        text = file_bytes.decode("utf-8", errors="replace")

    # Sniff delimiter
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t|;")
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    return [dict(row) for row in reader]


def _read_excel(file_bytes):
    """Parse Excel bytes → list of row dicts (uses openpyxl)."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl --break-system-packages")
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    # Find the header row — skip blank/logo rows at the top
    header_row_idx = 0
    for i, row in enumerate(rows):
        non_empty = [v for v in row if v is not None and str(v).strip()]
        if len(non_empty) >= 2:
            header_row_idx = i
            break
    headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(rows[header_row_idx])]
    result = []
    for row in rows[header_row_idx + 1:]:
        if all(v is None for v in row):
            continue
        result.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
    return result


# ── CSV / Excel import ───────────────────────────────────────────────────────

def _do_import(f, client_id=None, quarter_id=None, name=None, mapping=None):
    """
    Shared import logic. Returns (statement_id, transactions) or raises.
    mapping: optional dict like {"date": "Trans Date", "description": "Narrative",
             "amount": "Amount", "debit": "Debit", "credit": "Credit"}
    When mapping is provided, columns are read directly by mapped name.
    When mapping is None, the flexible alias system is used as fallback.
    """
    filename  = f.filename or "import"
    ext       = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    raw_bytes = f.read()

    if ext == "csv":
        raw_rows = _read_csv(raw_bytes, filename)
    elif ext in ("xlsx", "xls"):
        raw_rows = _read_excel(raw_bytes)
    else:
        raise ValueError(f"Unsupported file type: .{ext}. Upload a .csv or .xlsx file.")

    if not raw_rows:
        raise ValueError("File appears empty — no rows found.")

    if mapping:
        # Use explicit user-defined mapping
        def _mapped_row(raw, idx):
            date_col   = mapping.get("date", "")
            desc_col   = mapping.get("description", "")
            amt_col    = mapping.get("amount", "")
            debit_col  = mapping.get("debit", "")
            credit_col = mapping.get("credit", "")

            date  = str(raw.get(date_col, "")).strip()
            desc  = str(raw.get(desc_col, "")).strip()

            if debit_col or credit_col:
                debit_val  = _parse_money(raw.get(debit_col))
                credit_val = _parse_money(raw.get(credit_col))
                if credit_val and credit_val != 0:
                    amount = abs(credit_val)
                elif debit_val and debit_val != 0:
                    amount = -abs(debit_val)
                else:
                    amount = 0.0
            elif amt_col:
                amount = _parse_money(raw.get(amt_col)) or 0.0
            else:
                amount = 0.0

            return {"date": date, "description": desc, "amount": amount, "source_page": None}

        txns = [_mapped_row(r, i) for i, r in enumerate(raw_rows)]
    else:
        txns = [_normalise_row(r, i) for i, r in enumerate(raw_rows)]

    txns = [t for t in txns if t["description"] or t["amount"]]

    if not txns:
        raise ValueError(
            "No valid transactions found. "
            "Check column mapping — expected: Date + Description + (Debit/Credit or Amount). "
            f"Columns found: {list(raw_rows[0].keys()) if raw_rows else 'none'}"
        )

    # Persist to DB
    conn      = get_db()
    stmt_name = name or filename.rsplit(".", 1)[0]
    cur = conn.execute(
        "INSERT INTO statements (statement_name, bank_id, filename, status, quarter_id, created_at) "
        "VALUES (?,?,?,?,?,?)",
        (stmt_name, "csv-import", filename, "parsed",
         quarter_id, datetime.utcnow().isoformat())
    )
    sid = cur.lastrowid
    for t in txns:
        conn.execute(
            "INSERT INTO transactions (statement_id, date, description, amount, source_page) "
            "VALUES (?,?,?,?,?)",
            (sid, t["date"], t["description"], t["amount"], t["source_page"])
        )
    conn.commit()
    log_audit("statement", sid, "csv_import")

    rows = [dict(r) for r in conn.execute(
        "SELECT * FROM transactions WHERE statement_id = ? ORDER BY id", (sid,)
    ).fetchall()]

    return sid, rows



@ie_bp.route("/import/headers", methods=["POST"])
def get_csv_headers():
    """Read uploaded file and return its column headers for mapping UI."""
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded"}), 400
    ext = (f.filename or "").rsplit(".", 1)[-1].lower()
    raw_bytes = f.read()
    try:
        if ext == "csv":
            rows = _read_csv(raw_bytes, f.filename)
        elif ext in ("xlsx", "xls"):
            rows = _read_excel(raw_bytes)
        else:
            return jsonify({"error": f"Unsupported file type .{ext}"}), 400
        headers = list(rows[0].keys()) if rows else []
        # Return sample rows (first 3) so user can see data alongside headers
        sample = rows[:3] if rows else []
        return jsonify({"headers": headers, "sample": sample, "row_count": len(rows)})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


def _parse_mapping_from_request():
    """Extract column mapping from form data (sent as JSON string in 'mapping' field)."""
    import json
    mapping_raw = request.form.get("mapping")
    if mapping_raw:
        try:
            return json.loads(mapping_raw)
        except Exception:
            pass
    return None


@ie_bp.route("/import/csv", methods=["POST"])
def import_csv():
    """Original URL: /api/import/csv"""
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded"}), 400
    client_id  = request.form.get("client_id")
    quarter_id = request.form.get("quarter_id")
    name       = request.form.get("name")
    mapping    = _parse_mapping_from_request()
    try:
        sid, rows = _do_import(f, client_id, quarter_id, name, mapping)
        return jsonify({"statement_id": sid, "transactions": rows})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"Could not read file: {e}"}), 400


@ie_bp.route("/statements/import-csv", methods=["POST"])
def import_csv_alias():
    """JS-friendly alias: /api/statements/import-csv"""
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded"}), 400
    client_id  = request.form.get("client_id")
    quarter_id = request.form.get("quarter_id")
    name       = request.form.get("name")
    mapping    = _parse_mapping_from_request()
    try:
        sid, rows = _do_import(f, client_id, quarter_id, name, mapping)
        return jsonify({"statement_id": sid, "transactions": rows})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"Could not read file: {e}"}), 400


# ── GST Excel export ─────────────────────────────────────────────────────────

@ie_bp.route("/statements/<int:sid>/export/gst", methods=["GET"])
def export_gst(sid):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500
    try:
        conn = get_db()
        rows = [dict(r) for r in conn.execute(
            "SELECT t.*, c.name as category_name, c.pnl_group, c.bas_label, c.gst_applicable "
            "FROM transactions t LEFT JOIN categories c ON t.category_id = c.id "
            "WHERE t.statement_id = ?", (sid,)
        ).fetchall()]

        summary = gst_engine.summarize_gst(rows)
        bas    = summary.get("bas", {})
        by_cat = summary.get("by_category", [])

        wb  = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "BAS Summary"

        hdr_fill = PatternFill("solid", fgColor="1C3557")
        hdr_font = Font(color="FFFFFF", bold=True, size=11)
        bold     = Font(bold=True)
        center   = Alignment(horizontal="center")
        right    = Alignment(horizontal="right")

        ws1.append(["BAS Field", "Label", "Amount"])
        for cell in ws1[1]:
            cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center

        bas_fields = [
            ("G1",  "Total Sales (inc. GST)"),
            ("1A",  "GST on Sales"),
            ("G10", "Total Purchases (inc. GST)"),
            ("1B",  "GST Credits on Purchases"),
        ]
        for key, label in bas_fields:
            ws1.append([key, label, bas.get(key, 0)])
            ws1.cell(ws1.max_row, 3).number_format = "#,##0.00"
            ws1.cell(ws1.max_row, 3).alignment = right

        ws1.append([])
        net = (bas.get("1A", 0) or 0) - (bas.get("1B", 0) or 0)
        ws1.append(["NET GST", "Payable" if net >= 0 else "Refundable", net])
        r = ws1.max_row
        for c in [1, 2, 3]:
            ws1.cell(r, c).font = bold
        ws1.cell(r, 3).number_format = "#,##0.00"
        ws1.cell(r, 3).alignment = right
        ws1.column_dimensions["A"].width = 12
        ws1.column_dimensions["B"].width = 36
        ws1.column_dimensions["C"].width = 18

        ws2 = wb.create_sheet("By Category")
        ws2.append(["Category", "P&L Group", "BAS Label", "Gross (inc GST)", "GST Amount", "Net (ex GST)", "Txn Count"])
        for cell in ws2[1]:
            cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center
        for cat in by_cat:
            ws2.append([
                cat.get("category", ""), cat.get("pnl_group", ""),
                cat.get("bas_label", ""), cat.get("gross", 0),
                cat.get("gst", 0), cat.get("net", 0), cat.get("count", 0),
            ])
            for col in [4, 5, 6]:
                ws2.cell(ws2.max_row, col).number_format = "#,##0.00"
        for col in ["A","B","C","D","E","F","G"]:
            ws2.column_dimensions[col].width = 22

        ws3 = wb.create_sheet("Transactions")
        ws3.append(["Date", "Description", "Category", "BAS Label", "Amount (inc GST)", "GST (÷11)", "Net (ex GST)"])
        for cell in ws3[1]:
            cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center
        for t in rows:
            amt     = t.get("amount", 0) or 0
            has_gst = bool(t.get("gst_applicable"))
            gst_amt = round(amt / 11, 2) if has_gst else 0
            net_amt = round(amt - gst_amt, 2)
            ws3.append([
                t.get("date", ""), t.get("description", ""),
                t.get("category_name") or "Uncategorized",
                t.get("bas_label") or "",
                amt, gst_amt, net_amt,
            ])
            for col in [5, 6, 7]:
                ws3.cell(ws3.max_row, col).number_format = "#,##0.00"
        ws3.column_dimensions["A"].width = 14
        ws3.column_dimensions["B"].width = 50
        ws3.column_dimensions["C"].width = 24
        ws3.column_dimensions["D"].width = 14
        for col in ["E","F","G"]:
            ws3.column_dimensions[col].width = 18

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"gst_summary_stmt{sid}.xlsx")
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@ie_bp.route("/statements/<int:sid>/export/pnl", methods=["GET"])
def export_pnl(sid):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500
    try:
        conn = get_db()
        rows = [dict(r) for r in conn.execute(
            "SELECT t.*, c.name as category_name, c.pnl_group, c.bas_label, c.gst_applicable "
            "FROM transactions t LEFT JOIN categories c ON t.category_id = c.id "
            "WHERE t.statement_id = ?", (sid,)
        ).fetchall()]

        pnl = pnl_engine.generate_pnl(rows)

        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = "Profit & Loss"

        hdr_fill = PatternFill("solid", fgColor="1C3557")
        hdr_font = Font(color="FFFFFF", bold=True, size=11)
        bold     = Font(bold=True)
        bold13   = Font(bold=True, size=13)
        right    = Alignment(horizontal="right")
        grn_fill = PatternFill("solid", fgColor="D1FAE5")
        amb_fill = PatternFill("solid", fgColor="FEF3C7")
        red_fill = PatternFill("solid", fgColor="FEE2E2")
        sub_fill = PatternFill("solid", fgColor="EFF6FF")

        def section(title, lines, total, fill):
            if not lines:
                return
            ws.append([title, ""])
            for cell in ws[ws.max_row]:
                cell.font = hdr_font; cell.fill = hdr_fill
            for line in lines:
                ws.append(["  " + (line.get("category") or ""), line.get("amount", 0)])
                ws.cell(ws.max_row, 2).number_format = "#,##0.00"
                ws.cell(ws.max_row, 2).alignment = right
            ws.append(["Total " + title, total])
            r = ws.max_row
            ws.cell(r, 1).font = bold; ws.cell(r, 2).font = bold
            ws.cell(r, 2).number_format = "#,##0.00"
            ws.cell(r, 2).alignment = right
            for c in [1, 2]:
                ws.cell(r, c).fill = fill
            ws.append([])

        def subtotal_row(label, value, fill):
            ws.append([label, value])
            r = ws.max_row
            ws.cell(r, 1).font = bold13; ws.cell(r, 2).font = bold13
            ws.cell(r, 2).number_format = "#,##0.00"
            ws.cell(r, 2).alignment = right
            for c in [1, 2]:
                ws.cell(r, c).fill = fill
            ws.append([])

        section("Revenue",      pnl.get("income_lines",      []), pnl.get("total_income",      0), grn_fill)
        section("Direct Costs", pnl.get("direct_cost_lines", []), pnl.get("total_direct_cost", 0), amb_fill)
        if pnl.get("direct_cost_lines"):
            subtotal_row("Gross Profit", pnl.get("gross_profit", 0), sub_fill)
        section("Expenses", pnl.get("expense_lines", []), pnl.get("total_expense", 0), red_fill)

        net      = pnl.get("net_profit", 0)
        net_fill = PatternFill("solid", fgColor=("D1FAE5" if net >= 0 else "FEE2E2"))
        subtotal_row("Net Profit / Loss", net, net_fill)

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 20

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"pnl_stmt{sid}.xlsx")
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500
